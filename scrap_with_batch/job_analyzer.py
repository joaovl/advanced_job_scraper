#!/usr/bin/env python3
"""
Job Analyzer - Analyzes scraped jobs against CV using Ollama LLM
Produces Excel output with scoring and match analysis
"""

import json
import argparse
import logging
import requests
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import re

# Excel writing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class JobAnalyzer:
    def __init__(self, config_path: str = "config.json", cv_path: str = None):
        self.config = self.load_config(config_path)
        self.cv_content = self.load_cv(cv_path)
        self.ollama_url = self.config.get("ollama_url", "http://localhost:11434")
        self.ollama_model = self.config.get("ollama_model", "qwen2.5:latest")
        self.min_score = self.config.get("min_score", 7)
        self.exclude_in_title = [k.lower() for k in self.config.get("exclude_in_title", [])]
        self.exclude_in_description = [k.lower() for k in self.config.get("exclude_in_description", [])]
        self.flag_for_review = [k.lower() for k in self.config.get("flag_for_review", [])]
        self.must_have = [k.lower() for k in self.config.get("must_have", [])]

        # Load score adjustments for keyword weighting
        score_adj = self.config.get("score_adjustments", {})
        self.positive_keywords = {k.lower(): v for k, v in score_adj.get("positive", {}).items()}
        self.negative_keywords = {k.lower(): v for k, v in score_adj.get("negative", {}).items()}

    def load_config(self, config_path: str) -> Dict:
        """Load configuration from JSON file."""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Could not load config from {config_path}: {e}")
            return {}

    def load_cv(self, cv_path: str = None) -> str:
        """Load CV content from file."""
        # Try provided path first
        if cv_path:
            try:
                with open(cv_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    logger.info(f"Loaded CV from {cv_path} ({len(content)} chars)")
                    return content
            except Exception as e:
                logger.warning(f"Could not load CV from {cv_path}: {e}")

        # Try config path
        config_cv = self.config.get("cv_file", "")
        if config_cv:
            # Handle Docker path mapping
            local_paths = [
                config_cv,
                config_cv.replace("/data/n8n_linkedin/", ""),
                f"data/{Path(config_cv).name}",
                Path(config_cv).name
            ]
            for path in local_paths:
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        logger.info(f"Loaded CV from {path} ({len(content)} chars)")
                        return content
                except:
                    continue

        # Default fallback
        default_path = "data/your_cv.txt"
        try:
            with open(default_path, 'r', encoding='utf-8') as f:
                content = f.read()
                logger.info(f"Loaded CV from {default_path} ({len(content)} chars)")
                return content
        except Exception as e:
            logger.error(f"Could not load CV: {e}")
            return ""

    def load_jobs(self, jobs_path: str) -> List[Dict]:
        """Load jobs from JSON file."""
        try:
            with open(jobs_path, 'r', encoding='utf-8') as f:
                jobs = json.load(f)
                logger.info(f"Loaded {len(jobs)} jobs from {jobs_path}")
                return jobs
        except Exception as e:
            logger.error(f"Could not load jobs from {jobs_path}: {e}")
            return []

    def check_title_exclusion(self, title: str) -> Optional[str]:
        """Check if job title contains exclusion keywords (word boundary aware)."""
        title_lower = title.lower()
        for kw in self.exclude_in_title:
            kw_lower = kw.lower()
            # Use word boundary matching for short keywords (<=3 chars)
            if len(kw_lower) <= 3:
                # Match as whole word only
                pattern = r'\b' + re.escape(kw_lower) + r'\b'
                if re.search(pattern, title_lower):
                    return f"Title contains: {kw}"
            else:
                if kw_lower in title_lower:
                    return f"Title contains: {kw}"
        return None

    def check_description_exclusion(self, description: str) -> Optional[str]:
        """Check if job description contains exclusion keywords (word boundary aware)."""
        desc_lower = description.lower()
        for kw in self.exclude_in_description:
            kw_lower = kw.lower()
            # Use word boundary matching for short keywords (<=3 chars)
            if len(kw_lower) <= 3:
                pattern = r'\b' + re.escape(kw_lower) + r'\b'
                if re.search(pattern, desc_lower):
                    return f"Description contains: {kw}"
            else:
                if kw_lower in desc_lower:
                    return f"Description contains: {kw}"
        return None

    def check_flags(self, title: str, description: str) -> List[str]:
        """Check for review flags in title or description."""
        flags = []
        combined = (title + " " + description).lower()
        for kw in self.flag_for_review:
            if kw in combined:
                flags.append(kw)
        return flags

    def check_must_have(self, description: str) -> bool:
        """Check if description has at least one must-have keyword."""
        if not self.must_have:
            return True
        desc_lower = description.lower()
        for kw in self.must_have:
            if kw in desc_lower:
                return True
        return False

    def calculate_score_adjustment(self, title: str, description: str) -> tuple:
        """
        Calculate score adjustment based on weighted keywords.
        Returns (adjustment, matched_positive, matched_negative) tuple.
        """
        combined = (title + " " + description).lower()
        adjustment = 0
        matched_positive = []
        matched_negative = []

        # Check positive keywords
        for keyword, weight in self.positive_keywords.items():
            if keyword in combined:
                adjustment += weight
                matched_positive.append(f"{keyword}(+{weight})")

        # Check negative keywords
        for keyword, weight in self.negative_keywords.items():
            if keyword in combined:
                adjustment += weight  # weight is already negative
                matched_negative.append(f"{keyword}({weight})")

        return adjustment, matched_positive, matched_negative

    def call_ollama(self, job: Dict) -> Dict:
        """Call Ollama API to analyze job against CV."""
        cv_summary = self.cv_content[:2500] if self.cv_content else "No CV provided"

        system_prompt = """You are a job matching expert. Analyze if the job description matches the candidate CV. Return JSON only.

SCORING CRITERIA:
- 9-10: Perfect match - role aligns with experience, seniority, and domain
- 7-8: Good match - most requirements align, minor gaps acceptable
- 5-6: Partial match - some alignment but significant gaps
- 1-4: Poor match - wrong seniority, domain, or role type

REJECT if: Individual contributor coding role, hardware/electrical engineering, junior/entry level, non-software domain.

MATCH if: Engineering/QA management role, team leadership, test/release management, CI/CD, agile, software quality.

Return JSON: {"relevant": boolean, "rejection_reason": string (if not relevant), "score": number (1-10), "match_reasons": array of strings (top 3 reasons)}"""

        clean_desc = re.sub(r'[\x00-\x1F\x7F]', ' ', (job.get('description') or '')[:2000])
        clean_desc = re.sub(r'\s+', ' ', clean_desc)

        user_prompt = f"""CANDIDATE CV:
{cv_summary}

---

JOB TO ANALYZE:
Title: {job.get('title', 'Unknown')}
Company: {job.get('company', 'Unknown')}
Description: {clean_desc}

Analyze if this job matches the candidate. Return JSON only."""

        try:
            timeout = getattr(self, 'api_timeout', 180)
            response = requests.post(
                f"{self.ollama_url}/api/chat",
                json={
                    "model": self.ollama_model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    "stream": False,
                    "options": {"temperature": 0.3}
                },
                timeout=timeout
            )
            response.raise_for_status()

            result = response.json()
            content = result.get('message', {}).get('content', '')

            # Extract JSON from response
            content = re.sub(r'```json\n?', '', content, flags=re.IGNORECASE)
            content = re.sub(r'```\n?', '', content)
            content = content.strip()

            # Find JSON object
            match = re.search(r'\{[\s\S]*\}', content)
            if match:
                return json.loads(match.group(0))

            return {"relevant": False, "rejection_reason": "Could not parse response", "score": 0, "match_reasons": []}

        except requests.exceptions.RequestException as e:
            logger.error(f"Ollama API error: {e}")
            return {"relevant": False, "rejection_reason": f"API error: {str(e)[:50]}", "score": 0, "match_reasons": []}
        except json.JSONDecodeError as e:
            logger.error(f"JSON parse error: {e}")
            return {"relevant": False, "rejection_reason": "JSON parse error", "score": 0, "match_reasons": []}

    def analyze_job(self, job: Dict, index: int, total: int) -> Dict:
        """Analyze a single job and return result."""
        title = job.get('title', 'Unknown')
        company = job.get('company', 'Unknown')
        description = job.get('description', '')
        url = job.get('url', '')

        logger.info(f"[{index}/{total}] Analyzing: {title} at {company}")

        result = {
            'timestamp': datetime.now().isoformat(),
            'job_title': title,
            'company': company,
            'job_link': url,
            'location': job.get('location', ''),
            'decision': 'UNKNOWN',
            'reason': '',
            'score': 0,
            'ai_score': 0,
            'adjustment': 0,
            'score_details': '',
            'flags': ''
        }

        # Check for review flags
        flags = self.check_flags(title, description)
        if flags:
            result['flags'] = '; '.join(flags[:3])

        # Calculate score adjustment from keywords
        adjustment, matched_pos, matched_neg = self.calculate_score_adjustment(title, description)
        result['adjustment'] = adjustment
        if matched_pos or matched_neg:
            details = []
            if matched_pos:
                details.append(f"Boost: {', '.join(matched_pos[:5])}")
            if matched_neg:
                details.append(f"Reduce: {', '.join(matched_neg[:5])}")
            result['score_details'] = ' | '.join(details)

        # TIER 1: Title exclusion
        title_rejection = self.check_title_exclusion(title)
        if title_rejection:
            result['decision'] = 'REJECTED_TITLE'
            result['reason'] = title_rejection
            logger.info(f"  -> REJECTED (title): {title_rejection}")
            return result

        # TIER 2: Description exclusion
        desc_rejection = self.check_description_exclusion(description)
        if desc_rejection:
            result['decision'] = 'REJECTED_DESC'
            result['reason'] = desc_rejection
            logger.info(f"  -> REJECTED (desc): {desc_rejection}")
            return result

        # TIER 3: Must-have check
        if not self.check_must_have(description):
            result['decision'] = 'REJECTED_MISSING'
            result['reason'] = f"Missing required keywords: {', '.join(self.must_have[:3])}"
            logger.info(f"  -> REJECTED (missing must-have)")
            return result

        # TIER 4: No description - skip AI
        if not description or len(description.strip()) < 50:
            result['decision'] = 'SKIPPED'
            result['reason'] = 'No description available'
            logger.info(f"  -> SKIPPED (no description)")
            return result

        # TIER 5: AI Analysis
        logger.info(f"  -> Calling Ollama ({self.ollama_model})...")
        analysis = self.call_ollama(job)

        ai_score = int(analysis.get('score', 0))
        result['ai_score'] = ai_score

        # Apply keyword-based score adjustment
        final_score = ai_score + adjustment
        # Clamp score between 1 and 10
        final_score = max(1, min(10, final_score))
        result['score'] = final_score

        if adjustment != 0:
            logger.info(f"  -> AI score: {ai_score}, Adjustment: {adjustment:+d}, Final: {final_score}")

        if analysis.get('relevant') == True and final_score >= self.min_score:
            result['decision'] = 'MATCHED'
            reasons = analysis.get('match_reasons', [])[:3]
            result['reason'] = '; '.join(reasons) if reasons else 'Met criteria'
            logger.info(f"  -> MATCHED (final score {final_score})")
        else:
            result['decision'] = 'REJECTED_AI'
            result['reason'] = analysis.get('rejection_reason', f'Score {final_score} below threshold {self.min_score}')
            logger.info(f"  -> REJECTED (final score {final_score})")

        return result

    def analyze_all(self, jobs: List[Dict]) -> List[Dict]:
        """Analyze all jobs and return results."""
        results = []
        total = len(jobs)

        for i, job in enumerate(jobs, 1):
            result = self.analyze_job(job, i, total)
            results.append(result)

        return results

    def write_excel(self, results: List[Dict], output_path: str):
        """Write results to Excel file."""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not available, cannot write Excel")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Analysis"

        # Headers
        headers = ['Timestamp', 'Decision', 'Score', 'AI Score', 'Adjust', 'Score Details', 'Job Title', 'Company', 'Location', 'Reason', 'Flags', 'Job Link']

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        matched_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        rejected_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        skipped_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Write data
        for row_num, result in enumerate(results, 2):
            data = [
                result.get('timestamp', ''),
                result.get('decision', ''),
                result.get('score', ''),
                result.get('ai_score', ''),
                result.get('adjustment', ''),
                result.get('score_details', ''),
                result.get('job_title', ''),
                result.get('company', ''),
                result.get('location', ''),
                result.get('reason', ''),
                result.get('flags', ''),
                result.get('job_link', '')
            ]

            # Determine row color based on decision
            decision = result.get('decision', '')
            if decision == 'MATCHED':
                row_fill = matched_fill
            elif 'REJECTED' in decision:
                row_fill = rejected_fill
            elif decision == 'SKIPPED':
                row_fill = skipped_fill
            else:
                row_fill = None

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                if row_fill:
                    cell.fill = row_fill

                # Make job link clickable (column 12 now)
                if col == 12 and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        # Auto-adjust column widths
        column_widths = [20, 15, 8, 8, 8, 40, 40, 25, 20, 50, 30, 60]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Save
        wb.save(output_path)
        logger.info(f"Excel saved to {output_path}")

    def write_json(self, results: List[Dict], output_path: str):
        """Write results to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        logger.info(f"JSON saved to {output_path}")

    def print_summary(self, results: List[Dict]):
        """Print analysis summary."""
        matched = sum(1 for r in results if r['decision'] == 'MATCHED')
        rejected_title = sum(1 for r in results if r['decision'] == 'REJECTED_TITLE')
        rejected_desc = sum(1 for r in results if r['decision'] == 'REJECTED_DESC')
        rejected_ai = sum(1 for r in results if r['decision'] == 'REJECTED_AI')
        rejected_missing = sum(1 for r in results if r['decision'] == 'REJECTED_MISSING')
        skipped = sum(1 for r in results if r['decision'] == 'SKIPPED')

        print("\n" + "=" * 60)
        print("ANALYSIS SUMMARY")
        print("=" * 60)
        print(f"Total jobs analyzed: {len(results)}")
        print(f"  MATCHED:           {matched}")
        print(f"  REJECTED (title):  {rejected_title}")
        print(f"  REJECTED (desc):   {rejected_desc}")
        print(f"  REJECTED (AI):     {rejected_ai}")
        print(f"  REJECTED (missing):{rejected_missing}")
        print(f"  SKIPPED:           {skipped}")
        print("=" * 60)

        # Show matched jobs
        if matched > 0:
            print("\nMATCHED JOBS:")
            for r in results:
                if r['decision'] == 'MATCHED':
                    print(f"  [{r['score']}] {r['job_title']} at {r['company']}")
                    print(f"      {r['reason'][:80]}")
                    print(f"      {r['job_link']}")
        print()


def main():
    parser = argparse.ArgumentParser(description='Analyze jobs against CV using Ollama')
    parser.add_argument('jobs_file', help='Path to jobs JSON file')
    parser.add_argument('-c', '--config', default='config.json', help='Path to config file')
    parser.add_argument('--cv', help='Path to CV text file')
    parser.add_argument('-o', '--output', help='Output Excel file path')
    parser.add_argument('--json', help='Also save results as JSON to this path')
    parser.add_argument('--limit', type=int, help='Limit number of jobs to analyze')
    parser.add_argument('--company', help='Filter to specific company')
    parser.add_argument('--matched-only', action='store_true', help='Only show matched jobs in output')
    parser.add_argument('--model', help='Override Ollama model (e.g., qwen2.5:7b, llama3.1:8b)')
    parser.add_argument('--timeout', type=int, default=180, help='Ollama API timeout in seconds')
    parser.add_argument('--skip-analyzed', metavar='FILE', help='Skip jobs already in this analysis JSON file')
    parser.add_argument('--reanalyze', action='store_true', help='Force reanalysis even if in skip file (use with different model)')
    args = parser.parse_args()

    # Initialize analyzer
    analyzer = JobAnalyzer(config_path=args.config, cv_path=args.cv)

    # Override model if specified
    if args.model:
        analyzer.ollama_model = args.model
        logger.info(f"Using model: {args.model}")

    # Store timeout for API calls
    analyzer.api_timeout = args.timeout

    # Load previously analyzed jobs to skip
    analyzed_urls = set()
    if args.skip_analyzed and not args.reanalyze:
        try:
            with open(args.skip_analyzed, 'r', encoding='utf-8') as f:
                prev_results = json.load(f)
                analyzed_urls = {r.get('job_link', '') for r in prev_results if r.get('job_link')}
                logger.info(f"Loaded {len(analyzed_urls)} previously analyzed jobs from {args.skip_analyzed}")
        except Exception as e:
            logger.warning(f"Could not load skip file: {e}")

    # Load jobs
    jobs = analyzer.load_jobs(args.jobs_file)
    if not jobs:
        logger.error("No jobs to analyze")
        return

    # Filter by company if specified
    if args.company:
        jobs = [j for j in jobs if j.get('company', '').lower() == args.company.lower()]
        logger.info(f"Filtered to {len(jobs)} jobs for company: {args.company}")

    # Skip already analyzed jobs
    if analyzed_urls:
        original_count = len(jobs)
        jobs = [j for j in jobs if j.get('url', '') not in analyzed_urls]
        logger.info(f"Skipped {original_count - len(jobs)} already analyzed jobs, {len(jobs)} remaining")

    # Limit jobs if specified
    if args.limit:
        jobs = jobs[:args.limit]
        logger.info(f"Limited to {len(jobs)} jobs")

    # Analyze
    results = analyzer.analyze_all(jobs)

    # Filter matched only if requested
    if args.matched_only:
        results = [r for r in results if r['decision'] == 'MATCHED']

    # Generate output filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = args.output or f"job_analysis_{timestamp}.xlsx"

    # Write outputs
    analyzer.write_excel(results, output_path)

    # Always save JSON alongside Excel (for GUI loading)
    json_path = args.json or output_path.replace('.xlsx', '.json')
    analyzer.write_json(results, json_path)

    # Print summary
    analyzer.print_summary(results)

    print(f"\nResults saved to: {output_path}")


if __name__ == "__main__":
    main()
