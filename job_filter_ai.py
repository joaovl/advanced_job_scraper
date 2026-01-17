#!/usr/bin/env python3
"""
AI Job Filter - Uses Ollama to shortlist jobs matching your CV

This script reads scraped jobs and uses a local LLM (Ollama) to
score and filter them against your CV/preferences.

Usage:
    python job_filter_ai.py                     # Filter all jobs
    python job_filter_ai.py --limit 50          # Process first 50 jobs
    python job_filter_ai.py --location london   # Filter by location first
    python job_filter_ai.py --model llama3.2    # Use different model
"""

import json
import argparse
import requests
import sys
import subprocess
import os
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# For AI API calls - keep low to avoid rate limiting
# Claude API has rate limits, so don't overwhelm it
DEFAULT_AI_WORKERS = 4

# Check required dependencies upfront
def check_dependencies():
    """Check that all required Python packages are installed."""
    missing = []

    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")

    try:
        import requests
    except ImportError:
        missing.append("requests")

    if missing:
        print("=" * 60)
        print("ERROR: Missing required Python packages")
        print("=" * 60)
        print(f"\nMissing packages: {', '.join(missing)}")
        print("\nTo install, use one of these methods:\n")
        print("Option 1 - Using a virtual environment (recommended):")
        print("  python3 -m venv venv")
        print("  source venv/bin/activate  # On Linux/Mac")
        print("  venv\\Scripts\\activate     # On Windows")
        print(f"  pip install {' '.join(missing)}")
        print("\nOption 2 - Using system packages (Debian/Ubuntu):")
        apt_pkgs = ' '.join(f"python3-{pkg}" for pkg in missing)
        print(f"  sudo apt install {apt_pkgs}")
        print("\nOption 3 - Using pipx (for CLI tools):")
        print("  pipx install advanced-job-scraper  # if published")
        print()
        sys.exit(1)

check_dependencies()

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
N8N_DIR = BASE_DIR / "N8n"
OUTPUT_DIR = BASE_DIR / "output"

# Default config
DEFAULT_CONFIG = {
    "cv_file": str(N8N_DIR / "data" / "your_cv.txt"),
    "ollama_url": "http://localhost:11434",
    "ollama_model": "qwen3:8b",
    "use_llama_cli": False,
    "use_claude": False,
    "claude_model": "sonnet",  # sonnet, opus, or haiku
    "llama_cli_cmd": "llama-cli -hf Qwen/Qwen2.5-7B-Instruct-GGUF:Q4_K_M -ngl 99 -fa on -mg 1",
    "min_score": 7,
    "exclude_in_title": [
        "junior", "intern", "graduate", "entry level",
        "chef", "nurse", "teacher", "accountant",
        "sales", "marketing", "hr", "recruitment"
    ],
    "exclude_in_description": [
        "CSCS card", "construction", "manufacturing"
    ]
}


def load_config():
    """Load config from N8n/config.json or use defaults."""
    config_file = N8N_DIR / "config.json"
    if config_file.exists():
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # Merge with defaults
                for key, value in DEFAULT_CONFIG.items():
                    if key not in config:
                        config[key] = value
                return config
        except Exception as e:
            print(f"Warning: Could not load config.json: {e}")
    return DEFAULT_CONFIG.copy()


def load_cv(config: dict) -> str:
    """Load CV content from file."""
    cv_path = Path(config.get("cv_file", ""))

    # Try multiple locations
    paths_to_try = [
        cv_path,
        N8N_DIR / "data" / "your_cv.txt",
    ]

    for path in paths_to_try:
        if path.exists() and path.suffix == '.txt':
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return f.read()
            except:
                pass

    print("Warning: Could not load CV file. AI scoring will be less accurate.")
    return ""


def load_jobs(location_filter: str = None) -> list:
    """Load jobs from fintech_jobs.json or output directory."""
    # Try fintech_jobs.json first
    fintech_file = N8N_DIR / "fintech_jobs.json"
    if fintech_file.exists():
        with open(fintech_file, 'r', encoding='utf-8') as f:
            jobs = json.load(f)
        print(f"Loaded {len(jobs)} jobs from {fintech_file.name}")
    else:
        # Load from output directory
        print("fintech_jobs.json not found. Run 'python export_to_n8n.py' first.")
        return []

    # Apply location filter
    if location_filter:
        location_lower = location_filter.lower()
        jobs = [j for j in jobs if location_lower in j.get('location', '').lower()]
        print(f"Filtered to {len(jobs)} jobs in '{location_filter}'")

    return jobs


def check_claude() -> bool:
    """Check if Claude CLI is available."""
    try:
        result = subprocess.run(
            ["claude", "--version"],
            capture_output=True,
            text=True,
            timeout=10
        )
        if result.returncode == 0:
            print(f"Claude CLI version: {result.stdout.strip()}")
            return True
        return False
    except FileNotFoundError:
        print("Claude CLI not found. Install with: npm install -g @anthropic-ai/claude-code")
        return False
    except Exception as e:
        print(f"Claude CLI check error: {e}")
        return False


def check_llama_cli(cmd: str) -> bool:
    """Check if llama-cli is available."""
    try:
        # Just check if llama-cli exists
        result = subprocess.run(
            ["llama-cli", "--version"],
            capture_output=True,
            text=True,
            timeout=10
        )
        if result.returncode == 0 or "llama" in result.stdout.lower() or "llama" in result.stderr.lower():
            return True
        # Some versions don't have --version, try --help
        result = subprocess.run(
            ["llama-cli", "--help"],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.returncode == 0
    except FileNotFoundError:
        print("llama-cli not found. Make sure llama.cpp is installed and in PATH.")
        return False
    except Exception as e:
        print(f"llama-cli check error: {e}")
        return False


def check_ollama(url: str, model: str) -> bool:
    """Check if Ollama is running and model is available."""
    try:
        response = requests.get(f"{url}/api/tags", timeout=5)
        if response.status_code == 200:
            models = [m['name'] for m in response.json().get('models', [])]
            if model in models or any(model in m for m in models):
                return True
            print(f"Model '{model}' not found. Available: {', '.join(models[:5])}")
            return False
    except requests.exceptions.ConnectionError:
        print(f"Cannot connect to Ollama at {url}")
        print("Start Ollama with: ollama serve")
        return False
    except Exception as e:
        print(f"Ollama error: {e}")
        return False


def quick_filter(job: dict, exclude_title: list, exclude_desc: list) -> tuple:
    """Quick keyword-based filtering before AI."""
    import re
    title_lower = job.get('title', '').lower()
    desc_lower = job.get('description', '').lower()

    for kw in exclude_title:
        kw_lower = kw.lower()
        # Use word boundary matching for short keywords to avoid false positives
        if len(kw_lower) <= 3:
            if re.search(r'\b' + re.escape(kw_lower) + r'\b', title_lower):
                return False, f"Title contains: {kw}"
        else:
            if kw_lower in title_lower:
                return False, f"Title contains: {kw}"

    for kw in exclude_desc:
        kw_lower = kw.lower()
        if len(kw_lower) <= 3:
            if re.search(r'\b' + re.escape(kw_lower) + r'\b', desc_lower):
                return False, f"Description contains: {kw}"
        else:
            if kw_lower in desc_lower:
                return False, f"Description contains: {kw}"

    return True, ""


def score_job_with_ai(job: dict, cv: str, config: dict) -> dict:
    """Use Ollama, llama-cli, or Claude to score a job against the CV."""
    import re

    min_score = config.get('min_score', 7)

    prompt = f"""You are a job matching expert. Analyze if this job is a good match for the candidate.

SCORING (1-10):
- 9-10: Perfect match - role, seniority, and domain align well
- 7-8: Good match - most requirements align
- 5-6: Partial match - some alignment but gaps
- 1-4: Poor match - wrong level, domain, or role type

CANDIDATE CV:
{cv[:2000]}

JOB TO ANALYZE:
Title: {job.get('title', 'Unknown')}
Company: {job.get('company', 'Unknown')}
Location: {job.get('location', 'Unknown')}
Remote: {job.get('remote_type', 'Unknown')}
Description: {job.get('description', '')[:1500]}

Return ONLY a JSON object (no markdown, no explanation, just the JSON):
{{"score": <1-10>, "match": <true if score >= {min_score} else false>, "reasons": ["reason1", "reason2"]}}"""

    use_llama_cli = config.get('use_llama_cli', False)
    use_claude = config.get('use_claude', False)

    try:
        if use_claude:
            # Use Claude CLI
            result_text = score_with_claude(prompt, config)
        elif use_llama_cli:
            # Use llama-cli
            result_text = score_with_llama_cli(prompt, config)
        else:
            # Use Ollama API
            url = config.get('ollama_url', 'http://localhost:11434')
            model = config.get('ollama_model', 'llama3.2')

            response = requests.post(
                f"{url}/api/generate",
                json={
                    "model": model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.3}
                },
                timeout=120
            )
            response.raise_for_status()
            result_text = response.json().get('response', '')

        # Extract JSON from response
        match = re.search(r'\{[^{}]*\}', result_text)
        if match:
            result = json.loads(match.group())
            return {
                "score": int(result.get('score', 0)),
                "match": result.get('match', False),
                "reasons": result.get('reasons', [])
            }
    except Exception as e:
        return {"score": 0, "match": False, "reasons": [f"AI error: {str(e)}"]}

    return {"score": 0, "match": False, "reasons": ["Could not parse AI response"]}


def score_with_claude(prompt: str, config: dict, max_retries: int = 3) -> str:
    """Run Claude CLI with the given prompt and return the response. Includes retry logic."""
    import tempfile
    import time
    import random

    claude_model = config.get('claude_model', 'sonnet')

    for attempt in range(max_retries):
        # Write prompt to temp file with unique name to avoid conflicts
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
            f.write(prompt)
            prompt_file = f.name

        try:
            # Build claude command
            cmd = ["claude", "-p", prompt, "--output-format", "text"]

            # Add model if specified (haiku is cheaper/faster for this task)
            if claude_model in ['haiku', 'sonnet', 'opus']:
                cmd.extend(["--model", claude_model])

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=90,
                encoding='utf-8'
            )

            # Clean up temp file
            Path(prompt_file).unlink(missing_ok=True)

            if result.returncode != 0:
                error_msg = result.stderr[:200] if result.stderr else "Unknown error"
                # Check for rate limiting
                if "rate" in error_msg.lower() or "limit" in error_msg.lower() or "overloaded" in error_msg.lower():
                    if attempt < max_retries - 1:
                        wait_time = (attempt + 1) * 2 + random.uniform(0, 1)
                        time.sleep(wait_time)
                        continue
                raise Exception(f"Claude CLI error: {error_msg}")

            return result.stdout

        except subprocess.TimeoutExpired:
            Path(prompt_file).unlink(missing_ok=True)
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            raise Exception("Claude CLI timeout (90s)")
        except Exception as e:
            Path(prompt_file).unlink(missing_ok=True)
            if attempt < max_retries - 1 and ("rate" in str(e).lower() or "overloaded" in str(e).lower()):
                wait_time = (attempt + 1) * 2 + random.uniform(0, 1)
                time.sleep(wait_time)
                continue
            raise e

    raise Exception("Max retries exceeded")


def score_with_llama_cli(prompt: str, config: dict) -> str:
    """Run llama-cli with the given prompt and return the response."""
    import shlex
    import tempfile

    llama_cmd = config.get('llama_cli_cmd', 'llama-cli -hf Qwen/Qwen2.5-7B-Instruct-GGUF:Q4_K_M -ngl 99 -fa on -mg 1')

    # Write prompt to temp file to avoid shell escaping issues
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        f.write(prompt)
        prompt_file = f.name

    try:
        # Build command with prompt file
        cmd = f'{llama_cmd} -f "{prompt_file}" -n 256 --temp 0.3 -no-cnv'

        result = subprocess.run(
            cmd,
            shell=True,
            capture_output=True,
            text=True,
            timeout=180,
            encoding='utf-8'
        )

        # Clean up temp file
        Path(prompt_file).unlink(missing_ok=True)

        if result.returncode != 0:
            raise Exception(f"llama-cli error: {result.stderr[:200]}")

        return result.stdout

    except subprocess.TimeoutExpired:
        Path(prompt_file).unlink(missing_ok=True)
        raise Exception("llama-cli timeout (180s)")
    except Exception as e:
        Path(prompt_file).unlink(missing_ok=True)
        raise e


def filter_jobs(jobs: list, config: dict, cv: str, limit: int = None, parallel: bool = False, workers: int = None) -> list:
    """Filter jobs using quick filters and AI scoring."""
    exclude_title = config.get('exclude_in_title', [])
    exclude_desc = config.get('exclude_in_description', [])
    min_score = config.get('min_score', 7)

    results = []
    matched = 0
    rejected_quick = 0
    rejected_ai = 0

    jobs_to_process = jobs[:limit] if limit else jobs
    total = len(jobs_to_process)

    print(f"\nProcessing {total} jobs...")
    print("-" * 60)

    # Step 1: Quick filter (fast, sequential)
    jobs_for_ai = []
    for i, job in enumerate(jobs_to_process):
        title = job.get('title', 'Unknown')[:50]

        passed, reason = quick_filter(job, exclude_title, exclude_desc)

        if not passed:
            results.append({
                **job,
                "decision": "REJECTED",
                "score": 0,
                "reason": reason
            })
            rejected_quick += 1
            print(f"[{i+1}/{total}] SKIP: {title} - {reason}")
        else:
            jobs_for_ai.append((i, job))

    # Step 2: AI scoring
    if not jobs_for_ai:
        print("No jobs passed quick filter.")
    elif parallel and len(jobs_for_ai) > 1:
        # Parallel AI scoring with clean output
        num_workers = workers or DEFAULT_AI_WORKERS
        print(f"\nScoring {len(jobs_for_ai)} jobs with AI in PARALLEL ({num_workers} workers)...")

        def score_single_job(item):
            idx, job = item
            ai_result = score_job_with_ai(job, cv, config)
            return idx, job, ai_result

        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = {executor.submit(score_single_job, item): item for item in jobs_for_ai}
            completed = 0

            for future in as_completed(futures):
                completed += 1
                try:
                    idx, job, ai_result = future.result()
                    title = job.get('title', 'Unknown')[:40]
                    company = job.get('company', 'Unknown')[:15]

                    score = ai_result.get('score', 0)
                    is_match = ai_result.get('match', False) and score >= min_score
                    reasons = ai_result.get('reasons', [])

                    if is_match:
                        results.append({
                            **job,
                            "decision": "MATCHED",
                            "score": score,
                            "reason": "; ".join(reasons[:2])
                        })
                        matched += 1
                        print(f"[{completed:3}/{len(jobs_for_ai)}] MATCH  {score:2} {company:15} {title}")
                    else:
                        results.append({
                            **job,
                            "decision": "REJECTED",
                            "score": score,
                            "reason": "; ".join(reasons[:2])
                        })
                        rejected_ai += 1
                        print(f"[{completed:3}/{len(jobs_for_ai)}] REJECT {score:2} {company:15} {title}")

                except Exception as e:
                    idx, job = futures[future]
                    title = job.get('title', 'Unknown')[:40]
                    results.append({
                        **job,
                        "decision": "REJECTED",
                        "score": 0,
                        "reason": f"AI error: {e}"
                    })
                    rejected_ai += 1
                    print(f"[{completed:3}/{len(jobs_for_ai)}] ERROR  -- {title[:50]}")
    else:
        # Sequential AI scoring
        for i, (idx, job) in enumerate(jobs_for_ai):
            title = job.get('title', 'Unknown')[:50]
            company = job.get('company', 'Unknown')[:20]

            print(f"[{i+1}/{len(jobs_for_ai)}] AI: {title} @ {company}...", end=" ", flush=True)
            ai_result = score_job_with_ai(job, cv, config)

            score = ai_result.get('score', 0)
            is_match = ai_result.get('match', False) and score >= min_score
            reasons = ai_result.get('reasons', [])

            if is_match:
                results.append({
                    **job,
                    "decision": "MATCHED",
                    "score": score,
                    "reason": "; ".join(reasons[:2])
                })
                matched += 1
                print(f"MATCH (score {score})")
            else:
                results.append({
                    **job,
                    "decision": "REJECTED",
                    "score": score,
                    "reason": "; ".join(reasons[:2])
                })
                rejected_ai += 1
                print(f"REJECT (score {score})")

    print("-" * 60)
    print(f"Results: {matched} matched, {rejected_quick} quick-rejected, {rejected_ai} AI-rejected")

    # Sort by score descending
    results.sort(key=lambda x: (-1 if x['decision'] == 'MATCHED' else 0, -x.get('score', 0)))

    return results


def create_excel_report(results: list, output_file: Path):
    """Create comprehensive Excel report with all job details."""
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    reject_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers for all sheets
    headers = ["Score", "Decision", "Company", "Job Title", "Location", "Department",
               "Remote Type", "Description", "AI Reason", "Apply Link"]
    col_widths = [8, 12, 20, 45, 25, 20, 15, 80, 50, 12]

    def setup_sheet(ws, title, jobs_list):
        """Setup a worksheet with jobs data."""
        ws.title = title

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        ws.freeze_panes = "A2"

        # Add data rows
        for row_idx, job in enumerate(jobs_list, 2):
            decision = job.get('decision', 'UNKNOWN')
            score = job.get('score', 0)

            # Choose row color based on decision
            if decision == 'MATCHED':
                row_fill = match_fill
            elif score == 0:
                row_fill = skip_fill  # Quick-filtered
            else:
                row_fill = reject_fill

            # Score
            cell = ws.cell(row=row_idx, column=1, value=score)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_alignment

            # Decision
            cell = ws.cell(row=row_idx, column=2, value=decision)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_alignment

            # Company
            cell = ws.cell(row=row_idx, column=3, value=job.get('company', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # Job Title
            cell = ws.cell(row=row_idx, column=4, value=job.get('title', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # Location
            cell = ws.cell(row=row_idx, column=5, value=job.get('location', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # Department
            cell = ws.cell(row=row_idx, column=6, value=job.get('department', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # Remote Type
            cell = ws.cell(row=row_idx, column=7, value=job.get('remote_type', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_alignment

            # Description (truncate for Excel cell limit)
            desc = job.get('description', '')[:5000] if job.get('description') else ''
            cell = ws.cell(row=row_idx, column=8, value=desc)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # AI Reason
            cell = ws.cell(row=row_idx, column=9, value=job.get('reason', ''))
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment

            # Apply Link (clickable)
            url = job.get('url', '')
            cell = ws.cell(row=row_idx, column=10, value="Apply" if url else "")
            if url:
                cell.hyperlink = url
                cell.font = link_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_alignment

        # Set column widths
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        return len(jobs_list)

    # Sheet 1: All Jobs (sorted by score descending)
    ws_all = wb.active
    all_sorted = sorted(results, key=lambda x: (-x.get('score', 0), x.get('company', '')))
    count_all = setup_sheet(ws_all, "All Jobs", all_sorted)

    # Sheet 2: Matched Jobs Only
    ws_matched = wb.create_sheet()
    matched = [j for j in results if j.get('decision') == 'MATCHED']
    matched_sorted = sorted(matched, key=lambda x: (-x.get('score', 0), x.get('company', '')))
    count_matched = setup_sheet(ws_matched, "Matched", matched_sorted)

    # Sheet 3: Rejected by AI (had a score but didn't match)
    ws_rejected = wb.create_sheet()
    rejected_ai = [j for j in results if j.get('decision') == 'REJECTED' and j.get('score', 0) > 0]
    rejected_sorted = sorted(rejected_ai, key=lambda x: (-x.get('score', 0), x.get('company', '')))
    count_rejected = setup_sheet(ws_rejected, "AI Rejected", rejected_sorted)

    # Sheet 4: Quick-filtered (skipped, score=0)
    ws_skipped = wb.create_sheet()
    skipped = [j for j in results if j.get('score', 0) == 0]
    count_skipped = setup_sheet(ws_skipped, "Quick Filtered", skipped)

    # Sheet 5: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Category").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Count").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill

    summary_data = [
        ("Total Jobs Processed", count_all),
        ("Matched (Recommended)", count_matched),
        ("AI Rejected", count_rejected),
        ("Quick Filtered (Keywords)", count_skipped),
    ]

    for row_idx, (label, count) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=count)

    # Company breakdown
    ws_summary.cell(row=7, column=1, value="Jobs by Company").font = Font(bold=True)
    company_counts = {}
    for job in results:
        company = job.get('company', 'Unknown')
        if company not in company_counts:
            company_counts[company] = {'total': 0, 'matched': 0}
        company_counts[company]['total'] += 1
        if job.get('decision') == 'MATCHED':
            company_counts[company]['matched'] += 1

    ws_summary.cell(row=8, column=1, value="Company").font = header_font
    ws_summary.cell(row=8, column=1).fill = header_fill
    ws_summary.cell(row=8, column=2, value="Total").font = header_font
    ws_summary.cell(row=8, column=2).fill = header_fill
    ws_summary.cell(row=8, column=3, value="Matched").font = header_font
    ws_summary.cell(row=8, column=3).fill = header_fill

    for row_idx, (company, counts) in enumerate(sorted(company_counts.items()), 9):
        ws_summary.cell(row=row_idx, column=1, value=company)
        ws_summary.cell(row=row_idx, column=2, value=counts['total'])
        ws_summary.cell(row=row_idx, column=3, value=counts['matched'])

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 10

    # Save workbook
    wb.save(output_file)
    print(f"Excel report saved: {output_file}")
    return count_matched


def save_results(results: list, output_file: Path):
    """Save results to JSON, shortlist JSON, and Excel report."""
    # Save full results JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    # Create shortlist JSON (matched only)
    shortlist = [r for r in results if r['decision'] == 'MATCHED']
    shortlist_file = output_file.with_name(output_file.stem + '_shortlist.json')

    with open(shortlist_file, 'w', encoding='utf-8') as f:
        json.dump(shortlist, f, indent=2, ensure_ascii=False)

    # Create Excel report
    excel_file = output_file.with_suffix('.xlsx')
    create_excel_report(results, excel_file)

    print(f"\nSaved {len(results)} results to {output_file}")
    print(f"Saved {len(shortlist)} shortlisted jobs to {shortlist_file}")

    # Print shortlist summary
    if shortlist:
        print("\n" + "=" * 60)
        print("SHORTLISTED JOBS")
        print("=" * 60)
        for job in shortlist[:15]:
            print(f"\n[Score {job.get('score', 0)}] {job.get('title', 'Unknown')}")
            print(f"  Company: {job.get('company', 'Unknown')}")
            print(f"  Location: {job.get('location', 'Unknown')} ({job.get('remote_type', '')})")
            print(f"  URL: {job.get('url', '')}")
            if job.get('reason'):
                print(f"  Why: {job.get('reason', '')[:80]}")

        if len(shortlist) > 15:
            print(f"\n... and {len(shortlist) - 15} more jobs")


def main():
    parser = argparse.ArgumentParser(
        description="AI Job Filter - Score jobs against your CV using AI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Run with Claude in parallel (fastest)
    python job_filter_ai.py --claude --claude-model haiku -l London --parallel

    # Run with Claude (sequential)
    python job_filter_ai.py --claude --claude-model haiku --location London

    # Run with Ollama in parallel
    python job_filter_ai.py --ollama --model qwen2.5:7b -l London --parallel

    # Run with llama.cpp (local) - requires: llama-cli in PATH
    python job_filter_ai.py --llama-cli --location London

    # Parallel with custom worker count
    python job_filter_ai.py --claude -p -w 20 -l London
        """
    )
    parser.add_argument("--location", "-l", help="Filter by location first (e.g., 'London')")
    parser.add_argument("--limit", "-n", type=int, help="Limit number of jobs to process")
    parser.add_argument("--min-score", type=int, help="Minimum score to match (default: 6)")
    parser.add_argument("--parallel", "-p", action="store_true", help="Score jobs in parallel (faster)")
    parser.add_argument("--workers", "-w", type=int, default=DEFAULT_AI_WORKERS,
                        help=f"Number of parallel workers (default: {DEFAULT_AI_WORKERS})")

    # Backend selection (mutually exclusive)
    backend_group = parser.add_argument_group("AI Backend (choose one)")
    backend_group.add_argument("--claude", action="store_true",
                               help="Use Claude API (cloud) - fast, accurate")
    backend_group.add_argument("--ollama", action="store_true",
                               help="Use Ollama (local) - requires 'ollama serve'")
    backend_group.add_argument("--llama-cli", action="store_true",
                               help="Use llama.cpp CLI (local) - requires llama-cli in PATH")

    # Model options
    model_group = parser.add_argument_group("Model options")
    model_group.add_argument("--claude-model", choices=["haiku", "sonnet", "opus"], default="haiku",
                             help="Claude model (default: haiku - faster/cheaper)")
    model_group.add_argument("--model", "-m", default="qwen2.5:7b",
                             help="Ollama model (default: qwen2.5:7b)")
    model_group.add_argument("--llama-cmd",
                             help="Custom llama-cli command")
    model_group.add_argument("--url", default="http://localhost:11434",
                             help="Ollama URL (default: http://localhost:11434)")

    args = parser.parse_args()

    print("=" * 60)
    print("AI JOB FILTER")
    print("=" * 60)

    # Load config
    config = load_config()

    # Override with CLI args
    if args.model:
        config['ollama_model'] = args.model
    if args.min_score:
        config['min_score'] = args.min_score
    if args.url:
        config['ollama_url'] = args.url
    if args.llama_cmd:
        config['llama_cli_cmd'] = args.llama_cmd
    if args.claude_model:
        config['claude_model'] = args.claude_model

    # Determine backend (explicit flags take priority)
    if args.claude:
        config['use_claude'] = True
        config['use_llama_cli'] = False
    elif args.ollama:
        config['use_claude'] = False
        config['use_llama_cli'] = False
    elif args.llama_cli:
        config['use_llama_cli'] = True
        config['use_claude'] = False

    use_llama_cli = config.get('use_llama_cli', False)
    use_claude = config.get('use_claude', False)

    # Determine model name for output filename
    if use_claude:
        model_name = f"claude_{config.get('claude_model', 'haiku')}"
    elif use_llama_cli:
        model_name = "llama_cli"
    else:
        model_name = f"ollama_{config.get('ollama_model', 'qwen2.5').replace(':', '_').replace('/', '_')}"

    if use_claude:
        claude_model = config.get('claude_model', 'haiku')
        print(f"Backend: Claude CLI")
        print(f"Model: {claude_model}")
        print(f"Min score: {config['min_score']}")

        # Check Claude CLI
        if not check_claude():
            print("\nMake sure Claude CLI is installed and configured")
            sys.exit(1)

    elif use_llama_cli:
        llama_cmd = config.get('llama_cli_cmd', 'llama-cli -hf Qwen/Qwen2.5-7B-Instruct-GGUF:Q4_K_M -ngl 99 -fa on -mg 1')
        print(f"Backend: llama-cli")
        print(f"Command: {llama_cmd}")
        print(f"Min score: {config['min_score']}")

        # Check llama-cli
        if not check_llama_cli(llama_cmd):
            print("\nMake sure llama.cpp is installed and llama-cli is in your PATH")
            sys.exit(1)

        print("llama-cli: OK")
    else:
        print(f"Backend: Ollama")
        print(f"Model: {config['ollama_model']}")
        print(f"Min score: {config['min_score']}")

        # Check Ollama
        if not check_ollama(config['ollama_url'], config['ollama_model']):
            print("\nTo install a model: ollama pull llama3.2")
            sys.exit(1)

        print("Ollama: OK")

    # Load CV
    cv = load_cv(config)
    if cv:
        print(f"CV: Loaded ({len(cv)} chars)")

    # Load jobs
    jobs = load_jobs(location_filter=args.location)
    if not jobs:
        sys.exit(1)

    # Filter jobs
    results = filter_jobs(jobs, config, cv, limit=args.limit, parallel=args.parallel, workers=args.workers)

    # Save results with model name in filename for easy comparison
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = OUTPUT_DIR / f"ai_filtered_{model_name}_{timestamp}.json"
    save_results(results, output_file)

    print(f"\nTo compare with another model, run again with different backend:")
    print(f"  --claude --claude-model haiku   (cloud, fast)")
    print(f"  --ollama --model llama3.2       (local, Ollama)")
    print(f"  --llama-cli                     (local, llama.cpp)")


if __name__ == "__main__":
    main()
