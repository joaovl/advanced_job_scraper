#!/usr/bin/env python3
"""
Export all scraped jobs to a single Excel file with clickable links.

Usage:
    python export_to_excel.py
    python export_to_excel.py --output my_jobs.xlsx
"""

import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"


def load_all_jobs():
    """Load all jobs from JSON files in output directory."""
    all_jobs = []

    # Get the most recent file for each company
    company_files = {}

    for json_file in OUTPUT_DIR.glob("*.json"):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            company = data.get("company", "Unknown")
            scraped_at = data.get("scraped_at", "")

            # Keep track of most recent file per company
            if company not in company_files or scraped_at > company_files[company]["scraped_at"]:
                company_files[company] = {
                    "scraped_at": scraped_at,
                    "jobs": data.get("jobs", []),
                    "file": json_file.name
                }
        except Exception as e:
            print(f"Error loading {json_file}: {e}")

    # Collect all jobs from most recent files
    for company, info in sorted(company_files.items()):
        print(f"  {company}: {len(info['jobs'])} jobs (from {info['file']})")
        for job in info["jobs"]:
            job["company"] = company  # Ensure company is set
            all_jobs.append(job)

    return all_jobs


def create_excel(jobs, output_file):
    """Create Excel file with all jobs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Jobs"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    link_font = Font(color="0563C1", underline="single")
    cell_alignment = Alignment(vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Company colors for alternating groups
    company_colors = [
        "E2EFDA",  # Light green
        "DDEBF7",  # Light blue
        "FCE4D6",  # Light orange
        "E4DFEC",  # Light purple
        "FFF2CC",  # Light yellow
        "D9E1F2",  # Light indigo
    ]

    # Headers
    headers = ["Company", "Job Title", "Location", "Department", "Apply Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    # Sort jobs by company
    jobs_sorted = sorted(jobs, key=lambda x: x.get("company", "").lower())

    # Add jobs
    current_company = None
    color_idx = 0

    for row_idx, job in enumerate(jobs_sorted, 2):
        company = job.get("company", "Unknown")

        # Change color when company changes
        if company != current_company:
            current_company = company
            color_idx = (color_idx + 1) % len(company_colors)

        row_fill = PatternFill(start_color=company_colors[color_idx],
                               end_color=company_colors[color_idx],
                               fill_type="solid")

        # Company
        cell = ws.cell(row=row_idx, column=1, value=company)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Job Title
        cell = ws.cell(row=row_idx, column=2, value=job.get("title", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Location
        cell = ws.cell(row=row_idx, column=3, value=job.get("location", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Department
        cell = ws.cell(row=row_idx, column=4, value=job.get("department", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Apply Link (clickable)
        url = job.get("url", "")
        cell = ws.cell(row=row_idx, column=5, value="Apply Here")
        if url:
            cell.hyperlink = url
            cell.font = link_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions['A'].width = 15  # Company
    ws.column_dimensions['B'].width = 60  # Job Title
    ws.column_dimensions['C'].width = 25  # Location
    ws.column_dimensions['D'].width = 25  # Department
    ws.column_dimensions['E'].width = 12  # Apply Link

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Company").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Jobs Count").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill

    # Count jobs per company
    company_counts = {}
    for job in jobs:
        company = job.get("company", "Unknown")
        company_counts[company] = company_counts.get(company, 0) + 1

    for row_idx, (company, count) in enumerate(sorted(company_counts.items()), 2):
        ws_summary.cell(row=row_idx, column=1, value=company)
        ws_summary.cell(row=row_idx, column=2, value=count)

    # Total row
    total_row = len(company_counts) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(jobs)).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

    # Save
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")
    print(f"Total jobs: {len(jobs)}")


def main():
    parser = argparse.ArgumentParser(description="Export jobs to Excel")
    parser.add_argument("--output", "-o", default=None, help="Output Excel file")
    args = parser.parse_args()

    if args.output:
        output_file = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"all_jobs_{timestamp}.xlsx"

    print("Loading jobs from JSON files...")
    jobs = load_all_jobs()

    if not jobs:
        print("No jobs found!")
        return

    print(f"\nCreating Excel file with {len(jobs)} jobs...")
    create_excel(jobs, output_file)


if __name__ == "__main__":
    main()
