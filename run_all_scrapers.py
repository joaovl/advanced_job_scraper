#!/usr/bin/env python3
"""
Master Job Scraper - Runs all scrapers and generates combined output.

Usage:
    python run_all_scrapers.py                    # Run all scrapers
    python run_all_scrapers.py --playwright-only  # Only Playwright scrapers
    python run_all_scrapers.py --workday-only     # Only Workday scrapers
    python run_all_scrapers.py --export-only      # Only generate Excel/JSON from existing data
"""

import json
import argparse
import subprocess
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Installing openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
SCRAPERS_DIR = BASE_DIR / "scrapers"


def run_workday_scrapers(location="London"):
    """Run Workday API scrapers."""
    print("\n" + "=" * 70)
    print("RUNNING WORKDAY SCRAPERS")
    print("=" * 70)

    workday_script = SCRAPERS_DIR / "workday_scraper.py"
    if workday_script.exists():
        # Workday scraper uses --search for location filtering
        result = subprocess.run(
            [sys.executable, str(workday_script), "--all", "--search", location],
            cwd=str(BASE_DIR)
        )
        return result.returncode == 0
    else:
        print(f"Workday scraper not found: {workday_script}")
        return False


def run_playwright_scrapers(location="London"):
    """Run Playwright-based scrapers."""
    print("\n" + "=" * 70)
    print("RUNNING PLAYWRIGHT SCRAPERS (Cisco, Google, IBM, Apple, Meta, Amazon)")
    print("=" * 70)

    playwright_script = SCRAPERS_DIR / "playwright_scraper_v2.py"
    if playwright_script.exists():
        result = subprocess.run(
            [sys.executable, str(playwright_script), "--all", "--location", location],
            cwd=str(BASE_DIR)
        )
        return result.returncode == 0
    else:
        print(f"Playwright scraper not found: {playwright_script}")
        return False


def run_html_scrapers(location="London"):
    """Run HTML-based scrapers (Greenhouse, Lever, etc.)."""
    print("\n" + "=" * 70)
    print("RUNNING HTML SCRAPERS (Greenhouse, Lever, Ashby, etc.)")
    print("=" * 70)

    html_script = SCRAPERS_DIR / "run_html_scrapers.py"
    if html_script.exists():
        result = subprocess.run(
            [sys.executable, str(html_script)],
            cwd=str(BASE_DIR)
        )
        return result.returncode == 0
    else:
        print(f"HTML scraper not found: {html_script}")
    return True  # Not critical if missing


def load_all_jobs():
    """Load all jobs from JSON files, keeping only most recent per company."""
    all_jobs = []
    company_files = {}

    for json_file in OUTPUT_DIR.glob("*.json"):
        # Skip master files
        if json_file.name.startswith("master_") or json_file.name.startswith("all_jobs"):
            continue

        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            company = data.get("company", "Unknown")
            scraped_at = data.get("scraped_at", "")
            jobs = data.get("jobs", [])

            if not jobs:
                continue

            # Keep track of most recent file per company
            if company not in company_files or scraped_at > company_files[company]["scraped_at"]:
                company_files[company] = {
                    "scraped_at": scraped_at,
                    "jobs": jobs,
                    "file": json_file.name,
                    "location": data.get("location_searched", "")
                }
        except Exception as e:
            print(f"  Warning: Error loading {json_file.name}: {e}")

    # Collect all jobs
    for company, info in sorted(company_files.items()):
        print(f"  {company}: {len(info['jobs'])} jobs")
        for job in info["jobs"]:
            job["company"] = company
        all_jobs.extend(info["jobs"])

    return all_jobs, company_files


def create_master_json(jobs, company_files, output_file):
    """Create master JSON file with all jobs."""
    master_data = {
        "generated_at": datetime.now().isoformat(),
        "total_jobs": len(jobs),
        "total_companies": len(company_files),
        "companies": {
            company: {
                "job_count": len(info["jobs"]),
                "scraped_at": info["scraped_at"],
                "source_file": info["file"]
            }
            for company, info in sorted(company_files.items())
        },
        "jobs": jobs
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(master_data, f, indent=2, ensure_ascii=False)

    print(f"\nMaster JSON saved: {output_file}")
    return master_data


def create_excel(jobs, output_file):
    """Create Excel file with all jobs and clickable links."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Jobs"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    company_colors = ["E2EFDA", "DDEBF7", "FCE4D6", "E4DFEC", "FFF2CC", "D9E1F2"]

    # Headers
    headers = ["Company", "Job Title", "Location", "Department", "Apply Link"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.freeze_panes = "A2"

    # Sort and add jobs
    jobs_sorted = sorted(jobs, key=lambda x: x.get("company", "").lower())
    current_company = None
    color_idx = 0

    for row_idx, job in enumerate(jobs_sorted, 2):
        company = job.get("company", "Unknown")
        if company != current_company:
            current_company = company
            color_idx = (color_idx + 1) % len(company_colors)

        row_fill = PatternFill(start_color=company_colors[color_idx],
                               end_color=company_colors[color_idx], fill_type="solid")

        # Company
        cell = ws.cell(row=row_idx, column=1, value=company)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Job Title
        cell = ws.cell(row=row_idx, column=2, value=job.get("title", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Location
        cell = ws.cell(row=row_idx, column=3, value=job.get("location", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Department
        cell = ws.cell(row=row_idx, column=4, value=job.get("department", ""))
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

        # Apply Link
        url = job.get("url", "")
        cell = ws.cell(row=row_idx, column=5, value="Apply")
        if url:
            cell.hyperlink = url
            cell.font = link_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Company").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Jobs").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill

    company_counts = {}
    for job in jobs:
        company = job.get("company", "Unknown")
        company_counts[company] = company_counts.get(company, 0) + 1

    for row_idx, (company, count) in enumerate(sorted(company_counts.items()), 2):
        ws_summary.cell(row=row_idx, column=1, value=company)
        ws_summary.cell(row=row_idx, column=2, value=count)

    total_row = len(company_counts) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(jobs)).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 10

    wb.save(output_file)
    print(f"Excel saved: {output_file}")


def generate_exports():
    """Generate master JSON and Excel from existing data."""
    print("\n" + "=" * 70)
    print("GENERATING EXPORTS")
    print("=" * 70)

    print("\nLoading all jobs from JSON files...")
    jobs, company_files = load_all_jobs()

    if not jobs:
        print("No jobs found!")
        return None, None

    print(f"\nTotal: {len(jobs)} jobs from {len(company_files)} companies")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Master JSON
    json_file = OUTPUT_DIR / f"master_jobs_{timestamp}.json"
    create_master_json(jobs, company_files, json_file)

    # Excel
    excel_file = OUTPUT_DIR / f"all_jobs_{timestamp}.xlsx"
    create_excel(jobs, excel_file)

    return json_file, excel_file


def main():
    parser = argparse.ArgumentParser(description="Master Job Scraper")
    parser.add_argument("--location", "-l", default="London", help="Location to search")
    parser.add_argument("--playwright-only", action="store_true", help="Only run Playwright scrapers")
    parser.add_argument("--workday-only", action="store_true", help="Only run Workday scrapers")
    parser.add_argument("--export-only", action="store_true", help="Only generate exports from existing data")
    parser.add_argument("--no-export", action="store_true", help="Skip export generation")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)

    print("=" * 70)
    print("MASTER JOB SCRAPER")
    print(f"Location: {args.location}")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    if not args.export_only:
        # Run scrapers
        if args.playwright_only:
            run_playwright_scrapers(args.location)
        elif args.workday_only:
            run_workday_scrapers(args.location)
        else:
            # Run all
            run_workday_scrapers(args.location)
            run_html_scrapers(args.location)
            run_playwright_scrapers(args.location)

    if not args.no_export:
        # Generate combined outputs
        json_file, excel_file = generate_exports()

        if json_file and excel_file:
            print("\n" + "=" * 70)
            print("COMPLETE!")
            print("=" * 70)
            print(f"\nOutput files:")
            print(f"  Master JSON: {json_file}")
            print(f"  Excel:       {excel_file}")
            print(f"\nUse master JSON for AI analysis.")


if __name__ == "__main__":
    main()
